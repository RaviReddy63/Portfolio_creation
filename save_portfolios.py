import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# ==================== CONFIGURATION ====================

OUTPUT_DIR = 'output'

CURRENT_COLS = [
    'EMPLOYEE_NAME', 'MANAGER_NAME', 'DIRECTOR_NAME',
    'PORT_CODE', 'CG_ECN', 'CG_NAME', 'CG_BILLINGSTATE',
    'RC_ECN', 'RC_NAME', 'RC_BILLINGSTATE',
    'SEGMENT', 'RETAINED/REMOVED'
]

FUTURE_COLS = [
    'EMPLOYEE_NAME', 'MANAGER_NAME', 'DIRECTOR_NAME',
    'PORT_CODE', 'HH_ECN', 'HH_NAME', 'HH_BILLINGSTATE',
    'RC_ECN', 'RC_NAME', 'RC_BILLINGSTATE',
    'SEGMENT', 'RETAINED/REMOVED'
]

ADDED_COLS = [
    'EMPLOYEE_NAME', 'MANAGER_NAME', 'DIRECTOR_NAME',
    'PORT_CODE', 'HH_ECN', 'HH_NAME', 'HH_BILLINGSTATE',
    'RC_ECN', 'RC_NAME', 'RC_BILLINGSTATE',
    'SEGMENT'
]

REMOVED_COLS = [
    'EMPLOYEE_NAME', 'MANAGER_NAME', 'DIRECTOR_NAME',
    'PORT_CODE', 'CG_ECN', 'CG_NAME', 'CG_BILLINGSTATE',
    'RC_ECN', 'RC_NAME', 'RC_BILLINGSTATE',
    'SEGMENT', 'RC_ECN_TAG', 'RETAINED/REMOVED'
]

# ECN columns to format as integers
ECN_COLS = {'CG_ECN', 'HH_ECN', 'RC_ECN'}

HEADER_FILL  = PatternFill('solid', start_color='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT    = Font(name='Arial', size=10)
REMOVED_FILL = PatternFill('solid', start_color='FCE4D6')  # light red for removed sheet


# ==================== UTILITIES ====================

def sanitize(name):
    return re.sub(r'[\\/*?:"<>|]', '_', str(name)).strip()


def to_int_safe(val):
    """Convert ECN values to int, return as-is if not possible."""
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return val


def prep_ecn_cols(df, cols):
    """Convert ECN columns to integers where present."""
    df = df.copy()
    for col in cols:
        if col in df.columns and col in ECN_COLS:
            df[col] = df[col].apply(to_int_safe)
    return df


def auto_fit_columns(ws):
    """Auto-fit all column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4


def style_sheet(ws, num_data_rows, highlight_fill=None):
    """Apply header formatting and data font to all cells."""
    for cell in ws[1]:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=2, max_row=num_data_rows + 1):
        for cell in row:
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical='center')
            if highlight_fill:
                cell.fill = highlight_fill


def write_sheet(wb, sheet_name, df, columns, highlight_fill=None):
    """Write dataframe to worksheet with formatting."""
    ws = wb.create_sheet(title=sheet_name)

    # Ensure all required columns exist
    for col in columns:
        if col not in df.columns:
            df[col] = ''

    df = prep_ecn_cols(df, columns)

    ws.append(columns)
    if not df.empty:
        for row in dataframe_to_rows(df[columns], index=False, header=False):
            ws.append(row)

    style_sheet(ws, len(df), highlight_fill)
    auto_fit_columns(ws)
    return ws


def add_removed_dropdown(ws, num_data_rows):
    """Add REMOVED/RETAIN dropdown to RETAINED/REMOVED column in Removed sheet."""
    if num_data_rows == 0:
        return

    # Find RETAINED/REMOVED column index
    col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == 'RETAINED/REMOVED':
            col_idx = i
            break

    if col_idx is None:
        return

    col_letter = get_column_letter(col_idx)
    dv = DataValidation(
        type='list',
        formula1='"REMOVED,RETAIN"',
        allow_blank=False,
        showDropDown=False   # False = dropdown arrow visible
    )
    dv.sqref = f'{col_letter}2:{col_letter}{num_data_rows + 1}'
    ws.add_data_validation(dv)


# ==================== DELTA COMPUTATION ====================

def compute_retained_removed_flag(current_df, future_df, port_code):
    """
    Returns sets of RC_ECNs for a given port_code:
        retained_ecns : in both current and future
        removed_ecns  : in current only
        added_ecns    : in future only
    """
    curr_ecns = set(
        current_df[current_df['PORT_CODE'] == port_code]['RC_ECN'].dropna().unique()
    )
    fut_ecns = set(
        future_df[future_df['PORT_CODE'] == port_code]['RC_ECN'].dropna().unique()
    )
    return (
        curr_ecns & fut_ecns,   # retained
        curr_ecns - fut_ecns,   # removed
        fut_ecns  - curr_ecns,  # added
    )


def build_current_sheet_df(current_df, port_code, retained_ecns, removed_ecns):
    df = current_df[current_df['PORT_CODE'] == port_code].copy()
    def label(ecn):
        if ecn in retained_ecns:
            return 'RETAINED'
        elif ecn in removed_ecns:
            return 'REMOVED'
        return ''
    df['RETAINED/REMOVED'] = df['RC_ECN'].apply(label)
    return df


def build_future_sheet_df(future_df, port_code, retained_ecns, added_ecns):
    df = future_df[future_df['PORT_CODE'] == port_code].copy()
    def label(ecn):
        if ecn in retained_ecns:
            return 'RETAINED'
        elif ecn in added_ecns:
            return 'ADDED'
        return ''
    df['RETAINED/REMOVED'] = df['RC_ECN'].apply(label)
    return df


def build_added_sheet_df(future_df, port_code, added_ecns):
    df = future_df[future_df['PORT_CODE'] == port_code].copy()
    return df[df['RC_ECN'].isin(added_ecns)].copy()


def build_removed_sheet_df(current_df, port_code, removed_ecns):
    df = current_df[current_df['PORT_CODE'] == port_code].copy()
    df = df[df['RC_ECN'].isin(removed_ecns)].copy()
    df['RETAINED/REMOVED'] = 'REMOVED'
    return df


# ==================== WORKBOOK BUILDER ====================

def build_workbook(curr_df, fut_df, added_df, removed_df):
    """Build and return a workbook with all 4 sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(wb, 'Current Portfolio',   curr_df,    CURRENT_COLS)
    write_sheet(wb, 'Future Portfolio',    fut_df,     FUTURE_COLS)
    write_sheet(wb, 'New Customers Added', added_df,   ADDED_COLS)

    ws_removed = write_sheet(wb, 'Customers Removed', removed_df, REMOVED_COLS,
                              highlight_fill=REMOVED_FILL)
    add_removed_dropdown(ws_removed, len(removed_df))

    return wb


# ==================== MAIN ====================

def generate_portfolio_excels(current_portfolio_data, new_portfolio_data, portfolio_data):
    """
    Generate Excel files at three levels:
        1. Per banker  : output/Director/Manager/Banker/PORT_CODE.xlsx
        2. Per manager : output/Director/Manager/Manager_AGGREGATED.xlsx
        3. Per director: output/Director/Director_AGGREGATED.xlsx

    Args:
        current_portfolio_data : DataFrame — current portfolio state
        new_portfolio_data     : DataFrame — future/recommended portfolio state
        portfolio_data         : DataFrame — PORT_CODE, EMPLOYEE_NAME, MANAGER_NAME, DIRECTOR_NAME
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Rename RELATED_ECN -> RC_ECN if needed
    if 'RELATED_ECN' in new_portfolio_data.columns:
        new_portfolio_data = new_portfolio_data.rename(columns={'RELATED_ECN': 'RC_ECN'})

    # ---- Accumulate aggregated data ----
    # director_name -> {curr_rows, fut_rows, added_rows, removed_rows}
    director_agg = {}
    # (director_name, manager_name) -> {curr_rows, fut_rows, added_rows, removed_rows}
    manager_agg  = {}

    all_port_codes = portfolio_data['PORT_CODE'].dropna().unique()
    files_created  = 0

    for port_code in all_port_codes:
        banker_row = portfolio_data[portfolio_data['PORT_CODE'] == port_code]
        if banker_row.empty:
            continue

        banker_row    = banker_row.iloc[0]
        director_name = sanitize(banker_row['DIRECTOR_NAME'])
        manager_name  = sanitize(banker_row['MANAGER_NAME'])
        employee_name = sanitize(banker_row['EMPLOYEE_NAME'])

        # ---- Compute delta sets ----
        retained_ecns, removed_ecns, added_ecns = compute_retained_removed_flag(
            current_portfolio_data, new_portfolio_data, port_code
        )

        curr_df    = build_current_sheet_df(current_portfolio_data, port_code,
                                             retained_ecns, removed_ecns)
        fut_df     = build_future_sheet_df(new_portfolio_data, port_code,
                                            retained_ecns, added_ecns)
        added_df   = build_added_sheet_df(new_portfolio_data, port_code, added_ecns)
        removed_df = build_removed_sheet_df(current_portfolio_data, port_code, removed_ecns)

        # ---- Write banker-level file ----
        folder = os.path.join(OUTPUT_DIR, director_name, manager_name, employee_name)
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, f'{sanitize(port_code)}.xlsx')

        wb = build_workbook(curr_df, fut_df, added_df, removed_df)
        wb.save(file_path)
        files_created += 1
        print(f'  Created banker file : {file_path}')

        # ---- Accumulate for manager aggregation ----
        mgr_key = (director_name, manager_name)
        if mgr_key not in manager_agg:
            manager_agg[mgr_key] = {
                'curr': [], 'fut': [], 'added': [], 'removed': []
            }
        manager_agg[mgr_key]['curr'].append(curr_df)
        manager_agg[mgr_key]['fut'].append(fut_df)
        manager_agg[mgr_key]['added'].append(added_df)
        manager_agg[mgr_key]['removed'].append(removed_df)

        # ---- Accumulate for director aggregation ----
        if director_name not in director_agg:
            director_agg[director_name] = {
                'curr': [], 'fut': [], 'added': [], 'removed': []
            }
        director_agg[director_name]['curr'].append(curr_df)
        director_agg[director_name]['fut'].append(fut_df)
        director_agg[director_name]['added'].append(added_df)
        director_agg[director_name]['removed'].append(removed_df)

    # ---- Write manager-level aggregated files ----
    for (director_name, manager_name), data in manager_agg.items():
        curr_all    = pd.concat(data['curr'],    ignore_index=True) if data['curr']    else pd.DataFrame()
        fut_all     = pd.concat(data['fut'],     ignore_index=True) if data['fut']     else pd.DataFrame()
        added_all   = pd.concat(data['added'],   ignore_index=True) if data['added']   else pd.DataFrame()
        removed_all = pd.concat(data['removed'], ignore_index=True) if data['removed'] else pd.DataFrame()

        folder = os.path.join(OUTPUT_DIR, director_name, manager_name)
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, f'{manager_name}_AGGREGATED.xlsx')

        wb = build_workbook(curr_all, fut_all, added_all, removed_all)
        wb.save(file_path)
        files_created += 1
        print(f'  Created manager file: {file_path}')

    # ---- Write director-level aggregated files ----
    for director_name, data in director_agg.items():
        curr_all    = pd.concat(data['curr'],    ignore_index=True) if data['curr']    else pd.DataFrame()
        fut_all     = pd.concat(data['fut'],     ignore_index=True) if data['fut']     else pd.DataFrame()
        added_all   = pd.concat(data['added'],   ignore_index=True) if data['added']   else pd.DataFrame()
        removed_all = pd.concat(data['removed'], ignore_index=True) if data['removed'] else pd.DataFrame()

        folder = os.path.join(OUTPUT_DIR, director_name)
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, f'{director_name}_AGGREGATED.xlsx')

        wb = build_workbook(curr_all, fut_all, added_all, removed_all)
        wb.save(file_path)
        files_created += 1
        print(f'  Created director file: {file_path}')

    print(f'\nDone. {files_created} files created under ./{OUTPUT_DIR}/')


# ==================== USAGE ====================

# generate_portfolio_excels(
#     current_portfolio_data=CURRENT_PORTFOLIO_DATA,
#     new_portfolio_data=NEW_PORTFOLIO_DATA,
#     portfolio_data=SBRM_PORTFOLIO_DATA
# )
